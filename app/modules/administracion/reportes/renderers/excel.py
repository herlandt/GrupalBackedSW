"""Render de reportes a Excel (.xlsx) con openpyxl (capa RENDER, devuelve bytes)."""

from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.administracion.reportes.repository import (
    GananciasData,
    PagoFilaData,
    PagoPorEstudianteData,
)

_HEADER_FILL = PatternFill("solid", fgColor="0F172A")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _encabezado(ws: Worksheet, columnas: list[str]) -> None:
    ws.append(columnas)
    for celda in ws[1]:
        celda.fill = _HEADER_FILL
        celda.font = _HEADER_FONT


def _to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def ganancias_excel(data: GananciasData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ganancias"
    _encabezado(ws, ["Métrica", "Valor"])
    ws.append(["Total recaudado", float(data.total)])
    ws.append(["Moneda", data.moneda])
    ws.append(["Pagos completados", data.cantidad_pagos])
    return _to_bytes(wb)


def pagos_por_estudiante_excel(filas: Sequence[PagoPorEstudianteData]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos por estudiante"
    _encabezado(ws, ["Estudiante", "Email", "Total pagado", "# pagos"])
    for f in filas:
        ws.append([f.nombre, f.email, float(f.total_pagado), f.cantidad_pagos])
    return _to_bytes(wb)


def historial_usuario_excel(filas: Sequence[PagoFilaData]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"
    _encabezado(ws, ["Fecha", "Monto", "Moneda", "Estado"])
    for f in filas:
        ws.append([f.fecha.strftime("%Y-%m-%d %H:%M"), float(f.monto), f.moneda, f.estado])
    return _to_bytes(wb)
